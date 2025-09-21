import io
import json
import os
from typing import Dict, Any, List
from datetime import datetime

from PIL import Image, ImageOps, ImageEnhance, ImageFilter
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import google.generativeai as genai
except Exception:  # pragma: no cover
    genai = None


# --- Image helpers ---

def preprocess_image(image_bytes: bytes) -> Image.Image:
    """Lightweight preprocessing with tunable speed/quality via env.

    PREPROCESS_SCALE (float): scale factor (default 1.5)
    PREPROCESS_FAST (bool): if true, skip heavier filters (default true)
    """
    image = Image.open(io.BytesIO(image_bytes))
    image = image.convert("L")

    # Tunable scaling
    try:
        scale = float(os.getenv("PREPROCESS_SCALE", "1.5"))
    except Exception:
        scale = 1.5
    try:
        w, h = image.size
        if scale and scale != 1.0:
            image = image.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.LANCZOS)
    except Exception:
        pass

    fast = os.getenv("PREPROCESS_FAST", "true").lower() == "true"
    # Enhance contrast and sharpness (lighter when fast)
    contrast = 1.4 if fast else 1.8
    sharp = 1.2 if fast else 1.6
    image = ImageEnhance.Contrast(image).enhance(contrast)
    image = ImageOps.autocontrast(image)
    if not fast:
        image = image.filter(ImageFilter.MedianFilter(size=3))
    image = ImageEnhance.Sharpness(image).enhance(sharp)
    return image

def images_from_pdf(file_obj) -> List[Image.Image]:
    """Return PIL images from PDF pages with tunable DPI and page limit.

    GEMINI_PDF_DPI (int): rasterization DPI (default 200)
    GEMINI_MAX_PAGES (int): max pages to process (default 4)
    """
    data = file_obj.read()
    try:
        dpi = int(os.getenv("GEMINI_PDF_DPI", "200"))
    except Exception:
        dpi = 200
    try:
        max_pages = int(os.getenv("GEMINI_MAX_PAGES", "4"))
    except Exception:
        max_pages = 4

    doc = fitz.open(stream=data, filetype="pdf")
    images: List[Image.Image] = []
    try:
        for i, page in enumerate(doc):
            if max_pages > 0 and i >= max_pages:
                break
            pix = page.get_pixmap(dpi=dpi)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
    finally:
        doc.close()
    return images


# --- Structuring with Gemini Vision ---

def _get_gemini_model_name() -> str:
    # Default to 2.5 Flash for Vision; fallback set separately
    return os.getenv("GEMINI_MODEL", "gemini-2.5-flash")


def _make_model():
    """Create a GenerativeModel, falling back if the requested model isn't available."""
    name = _get_gemini_model_name()
    try:
        return genai.GenerativeModel(name)
    except Exception:
        fb = os.getenv("GEMINI_FALLBACK_MODEL", "gemini-2.5-flash-lite")
        return genai.GenerativeModel(fb)


def _get_gen_config() -> Dict[str, Any]:
    """Generation config tuned to reliably emit JSON while preserving characters."""
    try:
        temp = float(os.getenv("GEMINI_TEMPERATURE", "0.1"))
    except Exception:
        temp = 0.1
    try:
        mot = int(os.getenv("GEMINI_MAX_OUTPUT_TOKENS", "6000"))
    except Exception:
        mot = 6000
    return {
        "temperature": temp,
        "top_p": 0.8,
        "top_k": 20,
        "max_output_tokens": mot,
    }

def _image_to_part(img: Image.Image) -> Dict[str, Any]:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return {"mime_type": "image/png", "data": buf.getvalue()}


def structure_with_gemini_vision(images: List[Image.Image]) -> Dict[str, Any]:
    """Use Gemini Vision to directly extract tabular data from images."""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key or genai is None or not images:
        # Fallback to empty text structure
        return _fallback_structure("")

    genai.configure(api_key=api_key)
    model = _make_model()
    prompt = (
        "You are an expert multilingual document and table parser for banking PDFs and scans.\n"
        "The pages may contain English and Amharic (Ethiopic). Detect languages automatically.\n"
        "Goal: Extract text EXACTLY as it appears in the document without any corrections or modifications.\n\n"
        "Return STRICT JSON ONLY with this schema:\n"
        "{\n  \"tables\": [ { \"name\": string, \"headers\": [string], \"rows\": [[string]] } ]\n}\n\n"
        "CRITICAL INSTRUCTIONS - CHARACTER PRESERVATION:\n"
        "- NEVER autocorrect, fix, or modify any characters, words, or text\n"
        "- NEVER fix what appears to be OCR errors or typos - preserve them exactly\n"
        "- NEVER transliterate Amharic/Ethiopic characters to Latin script\n"
        "- NEVER normalize or standardize formatting - keep original spacing, punctuation\n"
        "- NEVER correct obvious mistakes like 0/O, 1/l, 5/S - transcribe exactly as shown\n"
        "- NEVER standardize dates or numbers - keep original format (e.g., 12/5/23, not 2023-05-12)\n"
        "- NEVER add missing punctuation or correct grammar\n"
        "- NEVER change case (uppercase/lowercase) from what is shown\n\n"
        "STRUCTURING RULES:\n"
        "- Copy each character, symbol, and space EXACTLY as it appears in the image\n"
        "- Preserve all original formatting, spacing, and line breaks\n"
        "- Maintain original table structure and cell alignment. If headers are missing, use Column1, Column2, ...\n"
        "- If you CANNOT confidently detect a table, return a single table named 'main' with headers ['text'] and rows = each visual line as a separate row, preserving order.\n"
        "- Output JSON only. No comments or markdown.\n"
    )

    # Cap images sent to the model for latency control
    try:
        max_images = int(os.getenv("GEMINI_MAX_IMAGES", "4"))
    except Exception:
        max_images = 4
    use_images = images[:max_images]

    parts: List[Any] = [prompt, "\n--- IMAGES START ---\n"]
    for img in use_images:
        parts.append(_image_to_part(img))
    parts.append("\n--- IMAGES END ---\n")

    # Generate and safely extract text, handling cases where response.text is unavailable
    try:
    # Apply permissive safety settings to avoid false-positive blocks on banking docs
        safety_settings = [
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUAL", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"}
        ]
        response = model.generate_content(parts, generation_config=_get_gen_config(), safety_settings=safety_settings)
        out = ""
        # Prefer the quick accessor but guard against exceptions
        try:
            out = (response.text or "").strip()
        except Exception:
            out = ""
        if not out:
            # Attempt to gather text from candidates/parts
            try:
                for cand in getattr(response, "candidates", []) or []:
                    content = getattr(cand, "content", None)
                    parts_list = getattr(content, "parts", None) if content else None
                    if not parts_list:
                        continue
                    buf: List[str] = []
                    for p in parts_list:
                        txt = getattr(p, "text", None)
                        if txt:
                            buf.append(txt)
                        elif isinstance(p, dict) and p.get("text"):
                            buf.append(str(p["text"]))
                    if buf:
                        out = "\n".join(buf)
                        break
            except Exception:
                out = ""
    except Exception:
        out = ""

    if not out:
        return _fallback_structure("")

    try:
        data = json.loads(_extract_json(out))
        if not isinstance(data, dict) or "tables" not in data:
            return _fallback_structure("")
        return data
    except Exception:
        return _fallback_structure("")


def _extract_json(s: str) -> str:
    # Remove fencing if present
    s = s.strip()
    if s.startswith("```"):
        s = s.strip("`\n")
        # Drop potential leading 'json' token
        if s.lstrip().startswith("json"):
            s = s.split("\n", 1)[1] if "\n" in s else "{}"
    return s


# --- PDF export fallback ---

def images_to_pdf(images: List[Image.Image]) -> bytes:
    """Create a simple multi-page PDF from one or more PIL images.
    Keeps original image fidelity (best for multilingual scripts like Amharic)."""
    if not images:
        return b""
    rgb_images: List[Image.Image] = []
    for im in images:
        if im.mode != "RGB":
            rgb_images.append(im.convert("RGB"))
        else:
            rgb_images.append(im)
    buf = io.BytesIO()
    first, rest = rgb_images[0], rgb_images[1:]
    first.save(buf, format="PDF", save_all=True, append_images=rest)
    return buf.getvalue()


def _fallback_structure(text: str) -> Dict[str, Any]:
    # Very naive: split lines, split on whitespace, first row as headers if looks tabular
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    rows = [ln.split() for ln in lines]
    headers: List[str]
    body: List[List[str]]
    if rows and all(len(r) == len(rows[0]) for r in rows[1:]):
        headers = [f"col{i+1}" for i in range(len(rows[0]))]
        body = rows
    else:
        headers = ["text"]
        body = [[ln] for ln in lines]

    return {
        "tables": [
            {"name": "main", "headers": headers, "rows": body}
        ]
    }


# --- Excel export ---

def to_excel(structured: Dict[str, Any]) -> bytes:
    wb = Workbook()
    # Remove default sheet
    ws0 = wb.active
    wb.remove(ws0)

    def _parse_cell(val: Any):
        # Try to coerce numbers (including 1,234.56) and ISO-like dates
        if isinstance(val, (int, float)):
            return val
        s = str(val).strip()
        # Try ISO date
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(s, fmt).date()
                return dt
            except Exception:
                pass
        # Remove currency symbols for numeric parse, keep as float
        try:
            # Replace common thousands separators
            num = s.replace(",", "").replace("ETB", "").replace("USD", "").strip()
            if num and (num.replace(".", "", 1).lstrip("-+").isdigit()):
                return float(num)
        except Exception:
            pass
        return s

    for table in structured.get("tables", []):
        title = table.get("name") or "Sheet"
        title = title[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title)
        headers = table.get("headers", [])
        rows = table.get("rows", [])
        if headers:
            ws.append(headers)
            # Bold header row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
            # Freeze header row
            ws.freeze_panes = "A2"
        for r in rows:
            parsed = [_parse_cell(c) if c is not None else "" for c in r]
            ws.append(parsed)

        # Basic column width adjustment based on max length in first 200 rows
        max_cols = max(len(headers), max((len(r) for r in rows), default=0))
        for col_idx in range(1, max_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), min_col=col_idx, max_col=col_idx), start=1):
                val = row[0].value
                if val is None:
                    continue
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
